import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

CATEGORY_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # Light orange
BOLD_FONT = Font(bold=True)
THICK_BORDER = Border(bottom=Side(style='thick'), top=Side(style='thick'), left=Side(style='thick'), right=Side(style='thick'))
CENTER = Alignment(horizontal='center', vertical='center')

def format_vfdb_matrix(matrix_csv, mapping_csv, excel_out):
    df = pd.read_csv(matrix_csv)
    mapping = pd.read_csv(mapping_csv)
    # Use only the short gene symbol for columns
    gene_order = mapping['gene'].tolist()
    cat2genes = mapping.groupby('category')['gene'].apply(list)
    # Metadata columns (remove 'First gene')
    meta_cols = ["GENOME", "Prophage", "Host genus", "KB", "GC%"]
    # Reorder columns: meta + gene columns
    df = df[meta_cols + gene_order]
    # Remove duplicate header row if present
    if (df.iloc[0].astype(str).tolist() == df.columns.tolist()):
        df = df.iloc[1:]
    # Prepare gene row
    gene_row = meta_cols + gene_order
    # Write to Excel
    with pd.ExcelWriter(excel_out, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Virulence genes', index=False, header=False, startrow=2)
        ws = writer.sheets['Virulence genes']
        # Write category header row with merged cells and improved formatting
        col_idx = 1
        # Metadata columns
        for m in meta_cols:
            cell = ws.cell(row=1, column=col_idx)
            cell.value = m
            cell.font = BOLD_FONT
            cell.alignment = CENTER
            cell.border = THICK_BORDER
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
            col_idx += 1
        # Gene columns by category (merged, bold, filled only for first cell)
        for cat, genes in cat2genes.items():
            start = col_idx
            end = col_idx + len(genes) - 1
            # Only set value, font, alignment, fill for the first cell
            cell = ws.cell(row=1, column=start)
            cell.value = cat
            cell.font = BOLD_FONT
            cell.alignment = CENTER
            cell.fill = CATEGORY_FILL
            cell.border = THICK_BORDER
            if end > start:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            # Only set border for the rest of the merged region
            for c in range(start, end + 1):
                if c != start:
                    ws.cell(row=1, column=c).border = THICK_BORDER
            col_idx = end + 1
        # Write gene header row (bold, centered) ONLY for gene columns (not metadata columns)
        for i, g in enumerate(gene_order):
            cell = ws.cell(row=2, column=len(meta_cols) + 1 + i)
            cell.value = g
            cell.font = BOLD_FONT
            cell.alignment = CENTER
            cell.border = THICK_BORDER
        # Do NOT set .value or formatting for metadata columns in row 2 (since those cells are merged vertically)
        ws.freeze_panes = ws.cell(row=3, column=len(meta_cols) + 1)
        # Autosize columns for readability
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max(10, max_length + 2)
    return excel_out 
